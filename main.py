from tkinter import *
import tkinter as tk
from tkinter import font
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from datetime import datetime
import random
root2 = tk.Tk()
root2.title("Library Management")
root2.geometry("400x600")
root2.resizable("False","False")
root2.iconbitmap("icon.ico")
wb = Workbook()
wb = load_workbook("library_data.xlsx")
wb_activate = wb.active
em = Workbook()
em = load_workbook("Employee.xlsx")
em_activate = em.active
mm = Workbook()
mm = load_workbook("Member.xlsx")
mm_activate = mm.active
lb = Workbook()
lb = load_workbook("logbook.xlsx")
lb_activate = lb.active
bg = PhotoImage(file="bg.png")
mainframe = Label(root2, image=bg)
mainframe.place(x=0, y=0, relwidth=1, relheight=1)
frame = Frame(root2, bg="#1D2025", width=400)
frame.place(x=0, y=0)
logframe = Frame(root2)
regframe = Frame(root2)
#Login Frame
    #widget
middleframe = Frame(logframe, bg="#1C2834")
lbl1 = Label(middleframe, text="Account", bg="#1C2834", fg="white", font=("High Tower Text", 30),)
lbl2 = Label(middleframe, text="Username:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
ent_user = Entry(middleframe, width=30)
lbl3 = Label(middleframe, text="Password", bg="#1C2834", fg="white", font=("High Tower Text", 12))
ent_pass = Entry(middleframe, width=30, show="●")
btn_login = Button(middleframe, text="Login", font=("High Tower Text", 12), width=10, command=lambda:Login())
expandlbl2 = Label(middleframe, bg="#1C2834")
    #layout
middleframe.pack(expand=True)
lbl1.grid(row=1, column=0,padx=60, pady=10)
lbl2.grid(row=2, column=0,padx=60, pady=5, sticky=W)
lbl3.grid(row=4, column=0,padx=60, pady=5, sticky=W)
ent_user.grid(row=3, column=0,padx=5, pady=5, ipady=3)
ent_pass.grid(row=5, column=0,padx=5, pady=5, ipady=3)
btn_login.grid(row=6, column=0,padx=10, pady=10)
expandlbl2.grid(row=7, column=0, padx=150, pady=99)
    #Fuction
def Login():
    global mFound
    global eFound
    hidden_pass = Entry(middleframe)
    u = ent_user.get()
    p = ent_pass.get()
    current_datetime = datetime.now()
    date_time_str = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
    row = (u, date_time_str)
    mFound = False
    eFound = False
    if u:
        for each_cell in range(2, (mm_activate.max_row)+1):
            if (u == mm_activate["F"+str(each_cell)].value):
                cell_address = str(each_cell)
                eFound = False
                mFound = True
                break
        for each_cell2 in range(2, (em_activate.max_row)+1):
            if (u == em_activate["G"+str(each_cell2)].value):
                cell_address2 = str(each_cell2)
                eFound = True
                mFound = False
                break
    else:
        messagebox.showerror("Login","Enter a Username")
    if mFound == True:
        if p:
            hidden_pass.insert(0,mm_activate['G'+cell_address].value)
            if p == hidden_pass.get():
                lb_activate.append(row)
                lb.save("logbook.xlsx")
                root2.iconify()
                newinterface()
            else:
                messagebox.showerror("Login","Incorrect Password")
        else:
            messagebox.showerror("Login","Enter a Password")
    if eFound == True:
        if p:
            hidden_pass.insert(0,em_activate['H'+cell_address2].value)
            if p == hidden_pass.get():
                lb_activate.append(row)
                lb.save("logbook.xlsx")
                root2.iconify()
                newinterface()
            else:
                messagebox.showerror("Login","Incorrect Password")
        else:
            messagebox.showerror("Login","Enter a Password")
#Register Frame
    #widget
middleframe = Frame(regframe, bg="#1C2834")
btnframe = Frame(middleframe, bg="#1C2834")
lblr_reg = Label(middleframe, text="Register", bg="#1C2834", fg="white", font=("High Tower Text", 30))
lblr_name = Label(middleframe, text="Name:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
entr_name = Entry(middleframe, width=30)
lblr_age = Label(middleframe, text="Age:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
entr_age = Entry(middleframe, width=30)
lblr_gender = Label(middleframe,text="Gender:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
sselect=StringVar()
chk_male=Radiobutton(middleframe, text="Male",value="Male", bg="#1C2834", fg="#F0C39A", font=("High Tower Text", 12),variable=sselect)
chk_female=Radiobutton(middleframe, text="Female",value="Female", bg="#1C2834", fg="#F0C39A", font=("High Tower Text", 12),variable=sselect)
lblr_cont = Label(middleframe, text="Phone:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
entr_cont = Entry(middleframe, width=30)
lblr_add = Label(middleframe, text="Address:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
entr_add = Text(middleframe, height= 5,width=23)
lblr_user = Label(middleframe, text="Username:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
entr_user = Entry(middleframe, width=30)
lblr_pass = Label(middleframe, text="Password:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
entr_pass = Entry(middleframe, width=30, show="●")
btnr_reglog = Button(middleframe, width=10, text="Register", font=("High Tower Text", 12), command=lambda:registermember())
btnr_regemp = Button(middleframe, width=10, text="Register", font=("High Tower Text", 12), command=lambda:registeremployee())
member_btn = Button(btnframe, width=10, text="Member", bg="#1C2834", fg="#F0C39A",border=1, font=("High Tower Text", 12), command=lambda:member())
memberclk_btn = Button(btnframe, width=10, text="Member", fg="#1C2834", bg="#F0C39A",border=0, font=("High Tower Text", 12), command=lambda:member())
employee_btn = Button(btnframe, width=10, text="Employee", bg="#1C2834", fg="#F0C39A",border=1, font=("High Tower Text", 12), command=lambda:employee())
employeeclk_btn = Button(btnframe, width=10, text="Employee", fg="#1C2834", bg="#F0C39A",border=0, font=("High Tower Text", 12), command=lambda:employee())
expandlbl1 = Label(middleframe, bg="#1C2834")
    #layout
middleframe.pack(expand=True)
lblr_reg.grid(row=0, column=0, columnspan=3, pady=10)
btnframe.grid(row=1, column=0, columnspan=3)
member_btn.grid(row=1, column=0, padx=26)
employee_btn.grid(row=1, column=1, padx=26)
lblr_name.grid(row=2, column=0,padx=10, sticky=E)
entr_name.grid(row=2, column=1,padx=5,columnspan=2, ipady=3)
lblr_age.grid(row=3, column=0,padx=10, sticky=E)
entr_age.grid(row=3, column=1,padx=5,columnspan=2, ipady=3)
lblr_gender.grid(row=4, column=0,padx=10, sticky=E)
chk_male.grid(row=4, column=1,padx=5)
chk_female.grid(row=4, column=2,padx=5)
lblr_cont.grid(row=6, column=0,padx=10, sticky=E)
entr_cont.grid(row=6, column=1,padx=5,columnspan=2, ipady=3)
lblr_add.grid(row=7, column=0,padx=10, sticky=E)
entr_add.grid(row=7, column=1,padx=5,columnspan=2, ipady=3)
lblr_user.grid(row=8, column=0,padx=10, sticky=E)
entr_user.grid(row=8, column=1,padx=5,columnspan=2, ipady=3)
lblr_pass.grid(row=9, column=0,padx=10, sticky=E)
entr_pass.grid(row=9, column=1,padx=5,columnspan=2, ipady=3)
expandlbl1.grid(row=15, column=0, columnspan=3, padx=150, pady=45)
    #Function
def member():
    member_btn.grid_forget()
    employeeclk_btn.grid_forget()
    btnr_regemp.grid_forget()
    btnr_reglog.grid(row=14, column=0, columnspan=3, pady=2)
    employee_btn.grid(row=1, column=1, padx=26)
    memberclk_btn.grid(row=1, column=0, padx=26)
    expandlbl1.grid_forget()
    expandlbl1.grid(row=15, column=0, columnspan=3, padx=150, pady=26)
def employee():
    memberclk_btn.grid_forget()
    employee_btn.grid_forget()
    btnr_reglog.grid_forget()
    btnr_regemp.grid(row=14, column=0, columnspan=3, pady=2)
    employeeclk_btn.grid(row=1, column=1, padx=26)
    member_btn.grid(row=1, column=0, padx=26)
    expandlbl1.grid_forget()
    expandlbl1.grid(row=15, column=0, columnspan=3, padx=150, pady=26)
def registermember():
    n = entr_name.get()
    a = entr_age.get()
    ad = entr_add.get("1.0", END)
    c = entr_cont.get()
    u = entr_user.get()
    p = entr_pass.get()
    s = sselect.get()
    if n and a and ad and c and u and p:
        if a.isdigit():
            if c.isdigit():
                lastrow = str(mm_activate.max_row + 1)
                mm_activate['A'+ lastrow] = n
                mm_activate['B'+ lastrow] = a
                mm_activate['C'+ lastrow] = c
                mm_activate['D'+ lastrow] = ad
                mm_activate['E'+ lastrow] = s
                mm_activate['F'+ lastrow] = u
                mm_activate['G'+ lastrow] = p
                mm_activate['H'+ lastrow] = "Register"
                mm.save('Member.xlsx')
                messagebox.showinfo("Register", "Sign Up Complete")
                clear()
            else:
                messagebox.showerror("Register", "Invalid Phone Input")
        else:
            messagebox.showerror("Register", "Invalid Age Input")
    else:
        messagebox.showerror("Register", "Registration Incomplete")      
def registeremployee():
    n = entr_name.get()
    a = entr_age.get()
    ad = entr_add.get("1.0", END)
    c = entr_cont.get()
    u = entr_user.get()
    p = entr_pass.get()
    s = sselect.get()
    if n and a and ad and c and u and p:
        if a.isdigit():
            if c.isdigit():
                lastrow = str(em_activate.max_row + 1)
                em_activate['B'+ lastrow] = n
                em_activate['C'+ lastrow] = a
                em_activate['D'+ lastrow] = c
                em_activate['E'+ lastrow] = ad
                em_activate['F'+ lastrow] = s
                em_activate['G'+ lastrow] = u
                em_activate['H'+ lastrow] = p
                em_activate['I'+ lastrow] = "Register"
                em.save('Employee.xlsx')
                messagebox.showinfo("Register", "Sign Up Complete")
                clear()
            else:
                messagebox.showerror("Register", "Invalid Phone Input")
        else:
            messagebox.showerror("Register", "Invalid Age Input")
    else:
        messagebox.showerror("Register", "Registration Incomplete") 
def clear():
    entr_cont.delete(0,END)
    entr_name.delete(0,END)
    entr_age.delete(0,END)
    entr_add.delete(1.0,END)
    entr_user.delete(0,END)
    entr_pass.delete(0,END)
#btnFrame
    #widget
btn_login = Button(frame, text="Sign In",width=8,borderwidth=0, bg="#1C2834", fg="#F0C39A", command=lambda:login_frame(), font=("High Tower Text", 15),)
btn_register = Button(frame, text="Sign Up",width=8,borderwidth=0, bg="#1C2834", fg="#F0C39A", command=lambda:register_frame(), font=("High Tower Text", 15),)
btnc_login = Button(frame, text="Sign In",width=8,borderwidth=0, fg="#1C2834", bg="#F0C39A", command=lambda:login_frame(), font=("High Tower Text", 15),)
btnc_register = Button(frame, text="Sign Up",width=8,borderwidth=0, fg="#1C2834", bg="#F0C39A", command=lambda:register_frame(), font=("High Tower Text", 15),)
des_frame = Frame(frame, bg="#F0C39A")
des_lbl = Label(des_frame, bg="#F0C39A")
    #Layout
btn_login.grid(row=1, column= 0, padx=50)
btn_register.grid(row=1, column= 1, padx=50)
des_frame.grid(row=2, column=0, columnspan=2)
des_lbl.pack(padx=200)
    #function
def login_frame():
    btnc_login.grid(row=1, column= 0, padx=50)
    btn_login.grid_forget()
    btn_register.grid(row=1, column= 1, padx=50)
    btnc_register.grid_forget()
    logframe.place(x=50, y=80)
    regframe.place_forget()
def register_frame():
    btnc_register.grid(row=1, column= 1, padx=50)
    btn_register.grid_forget()
    btn_login.grid(row=1, column= 0, padx=50)
    btnc_login.grid_forget()
    regframe.place(x=50, y=80)
    logframe.place_forget()
#===============================================================================================================================================================
def newinterface():
    Tl = Toplevel()
    Tl.title("Library Management")
    Tl.geometry("1100x700")
    Tl.geometry("+{}+{}".format(root2.winfo_x(), 0))
    Tl.resizable("False","False")
    #widget
    bg1 = PhotoImage(file="bg1.png")
    background = Label(Tl, image=bg1)
    background.place(x=0, y=0, relheight=1, relwidth=1)
    windowframe = Frame(Tl, bg="#1D2025")
    windowframe.place(x=0, y=0)
    #widget
    mhome_btn = Button(windowframe, text="Home", width=13, border=1, bg="#2F2E2C", fg="#F0C39A", font=("High Tower Text", 15), command=lambda:mhome())
    mlibrary_btn = Button(windowframe, text="Library", width=13, border=1, bg="#2F2E2C", fg="#F0C39A", font=("High Tower Text", 15), command=lambda:mlibrary())
    mprofile_btn = Button(windowframe, text="Profile", width=13, border=1, bg="#2F2E2C", fg="#F0C39A", font=("High Tower Text", 15), command=lambda:mprofileframe())
    mhelp_btn = Button(windowframe, text="Help", width=13, border=1, bg="#2F2E2C", fg="#F0C39A", font=("High Tower Text", 15), command=lambda:mhelpframe())
    mclkhome_btn = Button(windowframe, text="Home", bg= "#F0C39A",fg="#2F2E2C",width=13, border=1, font=("High Tower Text", 15), command=lambda:mhome())
    mclklibrary_btn = Button(windowframe, text="Library", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:mlibrary())
    mclkprofile = Button(windowframe, text="Profile", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:mprofileframe())
    mclkhelp = Button(windowframe, text="Help", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:mhelpframe())
    home_btn = Button(windowframe, text="Home", width=13, border=1, bg="#2F2E2C", fg="#F0C39A", font=("High Tower Text", 15), command=lambda:home())
    library_btn = Button(windowframe, text="Library", width=13, border=1, bg="#2F2E2C", fg="#F0C39A", font=("High Tower Text", 15), command=lambda:library())
    profile = Button(windowframe, text="Profile", width=13, border=1, bg="#2F2E2C", fg="#F0C39A", font=("High Tower Text", 15), command=lambda:profileframe())
    help = Button(windowframe, text="Help", width=13, border=1, bg="#2F2E2C", fg="#F0C39A", font=("High Tower Text", 15), command=lambda:helpframe())
    logbook_btn = Button(windowframe, text="Log Book", bg="#2F2E2C", fg="#F0C39A", width=13, border=1, font=("High Tower Text", 15), command=lambda:logbook())
    memberlist_btn = Button(windowframe, text="Members List", bg="#2F2E2C", fg="#F0C39A", width=13, border=1, font=("High Tower Text", 15), command=lambda:memberlist())
    bookmanagement_btn = Button(windowframe, text="Book Management", bg="#2F2E2C", fg="#F0C39A", width=13, border=1, font=("High Tower Text", 15), command=lambda:management())
    clkhome_btn = Button(windowframe, text="Home", bg= "#F0C39A",fg="#2F2E2C",width=13, border=1, font=("High Tower Text", 15), command=lambda:home())
    clklibrary_btn = Button(windowframe, text="Library", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:library())
    clkprofile = Button(windowframe, text="Profile", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:profileframe())
    clkhelp = Button(windowframe, text="Help", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:helpframe())
    clklogbook_btn = Button(windowframe, text="Log Book", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:logbook())
    clkmemberlist_btn = Button(windowframe, text="Members List", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:memberlist())
    clkbookmanagement_btn = Button(windowframe, text="Book Management", bg= "#F0C39A",fg="#2F2E2C", width=13, border=1, font=("High Tower Text", 15), command=lambda:management())
    lg = PhotoImage(file="logo.png")
    layout = Text(windowframe, width=13, border=0, bg="#1D2025")
    logom = Label(windowframe, image=lg, border=0)
    #search if member or employee
    #profile
    u = ent_user.get()
    for each_cell in range(2, (mm_activate.max_row)+1):
        if (u == mm_activate["F"+str(each_cell)].value):
            cell_address = str(each_cell)
            break
    for each_cell2 in range(2, (em_activate.max_row)+1):
        if (u == em_activate["G"+str(each_cell2)].value):
            cell_address2 = str(each_cell2)
            break
    if mFound == True:
        logom.grid(row=0, column=0)
        mclkhome_btn.grid(row=1, column=0, pady=20, ipady=10)
        mlibrary_btn.grid(row=2, column=0, pady=20, ipady=10)
        mprofile_btn.grid(row=3, column=0, pady=20, ipady=10)
        mhelp_btn.grid(row=4, column=0, pady=20, ipady=10)
        layout.grid(row=5, column=0, pady=10)
    if eFound == True:
        logom.grid(row=0, column=0)
        clkhome_btn.grid(row=1, column=0, pady=10)
        library_btn.grid(row=2, column=0, pady=10)
        logbook_btn.grid(row=3, column=0, pady=10)
        memberlist_btn.grid(row=5, column=0, pady=10)
        bookmanagement_btn.grid(row=6, column=0, pady=10)
        profile.grid(row=7, column=0, pady=10)
        help.grid(row=8, column=0, pady=10)
        layout.grid(row=20, column=0, pady=10) 
#================================================================================================================================
    #frames
    home_frame = Frame(Tl, bg="light gray")
    library_frame = Frame(Tl, bg="#1A1C22")
    profile_frame = Frame(Tl, bg="#1A1C22")
    help_frame = Frame(Tl, bg="#1A1C22")
    logbook_frame = Frame(Tl, bg="#1A1C22")
    memberlist_frame = Frame(Tl, bg="#1A1C22")
    bookmanagement_frame = Frame(Tl, bg="#1A1C22")
    home_frame.place(x=185, y=60)
    def mhome():
        library_frame.place_forget()
        profile_frame.place_forget()
        help_frame.place_forget()
        home_frame.place(x=185, y=60)
        mhome_btn.grid_forget()
        mclklibrary_btn.grid_forget()
        mclkprofile.grid_forget()
        mclkhelp.grid_forget()
        mclkhome_btn.grid(row=1, column=0, pady=20, ipady=10)
        mlibrary_btn.grid(row=2, column=0, pady=20, ipady=10)
        mprofile_btn.grid(row=3, column=0, pady=20, ipady=10)
        mhelp_btn.grid(row=4, column=0, pady=20, ipady=10)
    def mlibrary():
        library_frame.place(x=190, y=50)
        profile_frame.place_forget()
        help_frame.place_forget()
        home_frame.place_forget()
        mclkhome_btn.grid_forget()
        mlibrary_btn.grid_forget()
        mclkprofile.grid_forget()
        mclkhelp.grid_forget()
        mhome_btn.grid(row=1, column=0, pady=20, ipady=10)
        mclklibrary_btn.grid(row=2, column=0, pady=20, ipady=10)
        mprofile_btn.grid(row=3, column=0, pady=20, ipady=10)
        mhelp_btn.grid(row=4, column=0, pady=20, ipady=10)
    def mprofileframe():
        library_frame.place_forget()
        profile_frame.place(x=185, y=50)
        help_frame.place_forget()
        home_frame.place_forget()
        mclkhome_btn.grid_forget()
        mclklibrary_btn.grid_forget()
        mprofile_btn.grid_forget()
        mclkhelp.grid_forget()
        mhome_btn.grid(row=1, column=0, pady=20, ipady=10)
        mlibrary_btn.grid(row=2, column=0, pady=20, ipady=10)
        mclkprofile.grid(row=3, column=0, pady=20, ipady=10)
        mhelp_btn.grid(row=4, column=0, pady=20, ipady=10)
    def mhelpframe():
        library_frame.place_forget()
        profile_frame.place_forget()
        help_frame.place(x=190, y=20)
        home_frame.place_forget()
        mclkhome_btn.grid_forget()
        mclklibrary_btn.grid_forget()
        mclkprofile.grid_forget()
        mhelp_btn.grid_forget()
        mhome_btn.grid(row=1, column=0, pady=20, ipady=10)
        mlibrary_btn.grid(row=2, column=0, pady=20, ipady=10)
        mprofile_btn.grid(row=3, column=0, pady=20, ipady=10)
        mclkhelp.grid(row=4, column=0, pady=20, ipady=10)
    def home():
        library_frame.place_forget()
        profile_frame.place_forget()
        help_frame.place_forget()
        logbook_frame.place_forget()
        memberlist_frame.place_forget()
        bookmanagement_frame.place_forget()
        home_frame.place(x=185, y=60)
        home_btn.grid_forget()
        clklibrary_btn.grid_forget()
        clkprofile.grid_forget()
        clkhelp.grid_forget()
        clklogbook_btn.grid_forget()
        clkmemberlist_btn.grid_forget()
        clkbookmanagement_btn.grid_forget()
        clkhome_btn.grid(row=1, column=0, pady=10)
        library_btn.grid(row=2, column=0, pady=10)
        logbook_btn.grid(row=3, column=0, pady=10)
        memberlist_btn.grid(row=5, column=0, pady=10)
        bookmanagement_btn.grid(row=6, column=0, pady=10)
        profile.grid(row=7, column=0, pady=10)
        help.grid(row=8, column=0, pady=10)
    def library():
        home_frame.place_forget()
        profile_frame.place_forget()
        help_frame.place_forget()
        logbook_frame.place_forget()
        memberlist_frame.place_forget()
        bookmanagement_frame.place_forget()
        library_frame.place(x=190, y=50)
        clkhome_btn.grid_forget()
        library_btn.grid_forget()
        clkprofile.grid_forget()
        clkhelp.grid_forget()
        clklogbook_btn.grid_forget()
        clkmemberlist_btn.grid_forget()
        clkbookmanagement_btn.grid_forget()
        home_btn.grid(row=1, column=0, pady=10)
        clklibrary_btn.grid(row=2, column=0, pady=10)
        logbook_btn.grid(row=3, column=0, pady=10)
        memberlist_btn.grid(row=5, column=0, pady=10)
        bookmanagement_btn.grid(row=6, column=0, pady=10)
        profile.grid(row=7, column=0, pady=10)
        help.grid(row=8, column=0, pady=10)
    def logbook():
        home_frame.place_forget()
        profile_frame.place_forget()
        help_frame.place_forget()
        logbook_frame.place(x=350, y=80)
        memberlist_frame.place_forget()
        bookmanagement_frame.place_forget()
        library_frame.place_forget()
        clklibrary_btn.grid_forget()
        clkhome_btn.grid_forget()
        clkprofile.grid_forget()
        clkhelp.grid_forget()
        logbook_btn.grid_forget()
        clkmemberlist_btn.grid_forget()
        clkbookmanagement_btn.grid_forget()
        home_btn.grid(row=1, column=0, pady=10)
        library_btn.grid(row=2, column=0, pady=10)
        clklogbook_btn.grid(row=3, column=0, pady=10)
        memberlist_btn.grid(row=5, column=0, pady=10)
        bookmanagement_btn.grid(row=6, column=0, pady=10)
        profile.grid(row=7, column=0, pady=10)
        help.grid(row=8, column=0, pady=10)
    def memberlist():
        home_frame.place_forget()
        profile_frame.place_forget()
        help_frame.place_forget()
        logbook_frame.place_forget()
        memberlist_frame.place(x=190, y=50)
        bookmanagement_frame.place_forget()
        library_frame.place_forget()
        clklibrary_btn.grid_forget()
        clkhome_btn.grid_forget()
        clkprofile.grid_forget()
        clkhelp.grid_forget()
        clklogbook_btn.grid_forget()
        memberlist_btn.grid_forget()
        clkbookmanagement_btn.grid_forget()
        home_btn.grid(row=1, column=0, pady=10)
        library_btn.grid(row=2, column=0, pady=10)
        logbook_btn.grid(row=3, column=0, pady=10)
        clkmemberlist_btn.grid(row=5, column=0, pady=10)
        bookmanagement_btn.grid(row=6, column=0, pady=10)
        profile.grid(row=7, column=0, pady=10)
        help.grid(row=8, column=0, pady=10)
    def management():
        home_frame.place_forget()
        profile_frame.place_forget()
        help_frame.place_forget()
        logbook_frame.place_forget()
        memberlist_frame.place_forget()
        bookmanagement_frame.place(x=190, y=50)
        library_frame.place_forget()
        clklibrary_btn.grid_forget()
        clkhome_btn.grid_forget()
        clkprofile.grid_forget()
        clkhelp.grid_forget()
        clklogbook_btn.grid_forget()
        clkmemberlist_btn.grid_forget()
        bookmanagement_btn.grid_forget()
        home_btn.grid(row=1, column=0, pady=10)
        library_btn.grid(row=2, column=0, pady=10)
        logbook_btn.grid(row=3, column=0, pady=10)
        memberlist_btn.grid(row=5, column=0, pady=10)
        clkbookmanagement_btn.grid(row=6, column=0, pady=10)
        profile.grid(row=7, column=0, pady=10)
        help.grid(row=8, column=0, pady=10)
    def profileframe():
        library_frame.place_forget()
        home_frame.place_forget()
        help_frame.place_forget()
        logbook_frame.place_forget()
        memberlist_frame.place_forget()
        bookmanagement_frame.place_forget()
        profile_frame.place(x=185, y=50)
        clkhome_btn.grid_forget()
        clklibrary_btn.grid_forget()
        profile.grid_forget()
        clkhelp.grid_forget()
        clklogbook_btn.grid_forget()
        clkmemberlist_btn.grid_forget()
        clkbookmanagement_btn.grid_forget()
        home_btn.grid(row=1, column=0, pady=10)
        library_btn.grid(row=2, column=0, pady=10)
        logbook_btn.grid(row=3, column=0, pady=10)
        memberlist_btn.grid(row=5, column=0, pady=10)
        bookmanagement_btn.grid(row=6, column=0, pady=10)
        clkprofile.grid(row=7, column=0, pady=10)
        help.grid(row=8, column=0, pady=10)
    def helpframe():
        library_frame.place_forget()
        home_frame.place_forget()
        logbook_frame.place_forget()
        memberlist_frame.place_forget()
        bookmanagement_frame.place_forget()
        profile_frame.place_forget()
        help_frame.place(x=190, y=20)
        clkhome_btn.grid_forget()
        clklibrary_btn.grid_forget()
        clkprofile.grid_forget()
        help.grid_forget()
        clklogbook_btn.grid_forget()
        clkmemberlist_btn.grid_forget()
        clkbookmanagement_btn.grid_forget()
        home_btn.grid(row=1, column=0, pady=10)
        library_btn.grid(row=2, column=0, pady=10)
        logbook_btn.grid(row=3, column=0, pady=10)
        memberlist_btn.grid(row=5, column=0, pady=10)
        bookmanagement_btn.grid(row=6, column=0, pady=10)
        profile.grid(row=7, column=0, pady=10)
        clkhelp.grid(row=8, column=0, pady=10)
    #home_frame
    des = """
    Book Venture is an application designed to revolutionize your library experience, 
    making every visit an effortless and enjoyable adventure. 
    With Book Venture, we've reimagined the way you interact with your library, 
    creating a seamless bridge between technology and literature.
    """
    llg = PhotoImage(file="long logo.png")
    back = PhotoImage(file="back.png")
    logo = Label(home_frame, image=llg, border=0)
    description = Label(home_frame, text=des ,bg="#070706", fg="white", font=("High Tower Text", 19))
    logo.pack()
    description.place(x=10, y=320)
    #library page
    searchbtn = Button(library_frame, text="Search", bg="#F0C39A", width=10, border=1, font=("High Tower Text", 12), command=lambda:search_library())
    searchent = Entry(library_frame, width=50)
    backbtn = Button(library_frame, image=back, bg="#1A1C22", border=0, command=lambda:view_data())
    SomeFrame = Frame(library_frame)
    n = Entry(library_frame)
    def on_tree_select(event):
        selected_item = tv1.selection()
        if selected_item:
            book_data = tv1.item(selected_item)["values"]
            n.delete(0, tk.END)
            n.insert(0, book_data[0])
            for each_cell in range(2, (wb_activate.max_row)+1):
                if (n.get() == wb_activate["A"+str(each_cell)].value):
                    cell_address = str(each_cell)
                    title=f"Title:\t\t{book_data[0]}\nAuthor:\t\t{book_data[1]}\nPublishdate:\t{book_data[2]}\nDescription:\n{wb_activate['H'+cell_address].value}"
                    messagebox.showinfo(book_data[0]+" information", title)
                    break   
    def view_data():
        global tv1
        tv1 = ttk.Treeview(SomeFrame, columns=("Name", "Author", "Publish Date","Location","Status"), show="headings")
        tv1.heading("Name", text="Name", anchor=CENTER)
        tv1.heading("Author", text="Author", anchor=CENTER)
        tv1.heading("Publish Date", text="Publish Date", anchor=CENTER)
        tv1.heading("Location", text="Location", anchor=CENTER)
        tv1.heading("Status", text="Status", anchor=CENTER)
        for each_cell in range(2, (wb_activate.max_row)+1):
            tv1.insert(parent='', index="end", values=(wb_activate['A'+str(each_cell)].value, wb_activate['B'+str(each_cell)].value, wb_activate['C'+str(each_cell)].value, wb_activate['D'+str(each_cell)].value, wb_activate['E'+str(each_cell)].value))
        tv1.column("Name", width=250)
        tv1.column("Author", width=250)
        tv1.column("Publish Date", width=150)
        tv1.column("Location", width=100)
        tv1.column("Status", width=100)
        tv1.bind("<<TreeviewSelect>>", on_tree_select)
        tree_scroll = ttk.Scrollbar(SomeFrame, orient="vertical",command=tv1.yview)
        tv1.configure(yscrollcommand=tree_scroll.set)
        tv1.config(height=25)
        tv1.grid(row=0, column=0, sticky="nsew")
        tree_scroll.grid(row=0, column=1, sticky="ns")
    view_data()
    def search_library():
        query = searchent.get().strip().lower()
        tv1.delete(*tv1.get_children())
        for row in wb_activate.iter_rows(min_row=2, values_only=True):
            if any(query in str(cell).lower() for cell in row):
                tv1.insert("", tk.END, values=row)     
    searchbtn.grid(row=1, column=1, pady=10)
    searchent.grid(row=1, column=0, sticky=E, padx=10, ipady=3)
    backbtn.grid(row=1, column=2, sticky=W)
    SomeFrame.grid(row=2, column=0, columnspan=3, padx=10, pady=10)
    #logbook page
    lsearchbtn = Button(logbook_frame, text="Search", bg="#F0C39A", width=10, border=1, font=("High Tower Text", 12), command=lambda:lsearch_library())
    lsearchent = Entry(logbook_frame, width=50)
    lbackbtn = Button(logbook_frame, image=back, bg="#1A1C22", border=0, command=lambda:lview_data())
    lSomeFrame = Frame(logbook_frame)
    def lview_data():
        global tv3
        tv3 = ttk.Treeview(lSomeFrame, columns=("User", "Date & Time"), show="headings")
        tv3.heading("User", text="User", anchor=CENTER)
        tv3.heading("Date & Time", text="Date & Time", anchor=CENTER)
        for each_cell in range(2, (lb_activate.max_row)+1):
            tv3.insert(parent='', index="end", values=(lb_activate['A'+str(each_cell)].value, lb_activate['B'+str(each_cell)].value))
        tv3.column("User", width=200)
        tv3.column("Date & Time", width=250)
        ltree_scroll = ttk.Scrollbar(lSomeFrame, orient="vertical",command=tv3.yview)
        tv3.configure(yscrollcommand=ltree_scroll.set)
        tv3.config(height=20)
        tv3.grid(row=0, column=0, sticky="nsew")
        ltree_scroll.grid(row=0, column=1, sticky="ns")
    lview_data()
    def lsearch_library():
        query = lsearchent.get().strip().lower()
        tv3.delete(*tv3.get_children())
        for row in lb_activate.iter_rows(min_row=2, values_only=True):
            if any(query in str(cell).lower() for cell in row):
                tv3.insert("", tk.END, values=row)
    lsearchbtn.grid(row=0, column=0, sticky=E, padx=10)
    lsearchent.grid(row=0, column=1, pady=10, ipady=3)
    lbackbtn.grid(row=0, column=2, sticky=W, padx=10)
    lSomeFrame.grid(row=1, column=0, columnspan=3, padx=10, pady=10)
    #memberlist page
    msearchbtn = Button(memberlist_frame, text="Search", bg="#F0C39A", width=10, border=1, font=("High Tower Text", 12), command=lambda:msearch_library())
    mbackbtn = Button(memberlist_frame, image=back, bg="#1A1C22", border=0, command=lambda:mview_data())
    msearchent = Entry(memberlist_frame, width=50)
    mSomeFrame = Frame(memberlist_frame)
    def mview_data():
        global tv2
        tv2 = ttk.Treeview(mSomeFrame, columns=("Name", "Age","Phone","Address","Sex"), show="headings")
        tv2.heading("Name", text="Name", anchor=CENTER)
        tv2.heading("Age", text="Age", anchor=CENTER)
        tv2.heading("Phone", text="Phone", anchor=CENTER)
        tv2.heading("Address", text="Address", anchor=CENTER)
        tv2.heading("Sex", text="Sex", anchor=CENTER)
        for each_cell in range(2, (mm_activate.max_row)+1):
            tv2.insert(parent='', index="end", values=(mm_activate['A'+str(each_cell)].value, mm_activate['B'+str(each_cell)].value, mm_activate['C'+str(each_cell)].value, mm_activate['D'+str(each_cell)].value, mm_activate['E'+str(each_cell)].value))
        tv2.column("Name", width=250)
        tv2.column("Age", width=50)
        tv2.column("Phone", width=100)
        tv2.column("Address", width=300)
        tv2.column("Sex", width=80)
        tree_scrolly = ttk.Scrollbar(mSomeFrame, orient="vertical",command=tv2.yview)
        tv2.configure(yscrollcommand=tree_scrolly.set)
        tree_scrollx = ttk.Scrollbar(mSomeFrame, orient="horizontal",command=tv2.xview)
        tv2.configure(xscrollcommand=tree_scrollx.set)
        tv2.config(height=25)
        tv2.grid(row=0, column=0, sticky="nsew")
        tree_scrolly.grid(row=0, column=1, sticky="ns")
        tree_scrollx.grid(row=1, column=0, sticky="ew")
    mview_data()
    def msearch_library():
        query = msearchent.get().strip().lower()
        tv2.delete(*tv2.get_children())
        for row in mm_activate.iter_rows(min_row=2, values_only=True):
            if any(query in str(cell).lower() for cell in row):
                tv2.insert("", tk.END, values=row)
    msearchbtn.grid(row=1, column=1, pady=10)
    msearchent.grid(row=1, column=0, sticky=E, padx=10, ipady=3)
    mbackbtn.grid(row=1, column=2, sticky=W)
    mSomeFrame.grid(row=2, column=0, columnspan=3, padx=10, pady=10)
    #management page
    name_lbl = Label(bookmanagement_frame, text="Name:", bg="#1A1C22", fg="white", font=("High Tower Text", 12))
    author_lbl = Label(bookmanagement_frame, text="Author:", bg="#1A1C22", fg="white", font=("High Tower Text", 12))
    publishdate_lbl = Label(bookmanagement_frame, text="Publish Date:", bg="#1A1C22", fg="white", font=("High Tower Text", 12))
    location_lbl = Label(bookmanagement_frame, text="Location:", bg="#1A1C22", fg="white", font=("High Tower Text", 12))
    name_ent = Entry(bookmanagement_frame, width=40)
    author_ent = Entry(bookmanagement_frame, width=40)
    publishdate_ent = Entry(bookmanagement_frame, width=40)
    location_ent = Entry(bookmanagement_frame, width=40)
    save_btn = Button(bookmanagement_frame, text="Save", width=10, height=2, border=1, bg="#F0C39A", fg="#1A1C22", font=("High Tower Text", 12), command=lambda:save())
    Remove_btn = Button(bookmanagement_frame, text="Remove", width=10, height=2, border=1, bg="#F0C39A", fg="#1A1C22", font=("High Tower Text", 12), command=lambda:remove())
    Edit_btn = Button(bookmanagement_frame, text="Edit", width=10, height=2, border=1, bg="#F0C39A", fg="#1A1C22", font=("High Tower Text", 12), command=lambda:edit())
    clear_btn = Button(bookmanagement_frame, text="Clear", width=10, height=2, border=1, bg="#F0C39A", fg="#1A1C22", font=("High Tower Text", 12), command=lambda:clear())
    search_ent = Entry(bookmanagement_frame, width=40)
    search_btn = Button(bookmanagement_frame, text="Search", width=10, height=2, border=1, bg="#F0C39A", fg="#1A1C22", font=("High Tower Text", 12), command=lambda:lmsearch_library())
    back_btn = Button(bookmanagement_frame, text="Back", width=10, height=2, border=1, bg="#F0C39A", fg="#1A1C22", font=("High Tower Text", 12), command=lambda:lmview_data())
    lmSomeFrame = Frame(bookmanagement_frame)
    code = Entry(bookmanagement_frame)
    def save():
        alreadyexist = False
        n = name_ent.get()
        a = author_ent.get()
        p = publishdate_ent.get()
        l = location_ent.get()
        if n and a and p and l:
            for every_row in range(2, (wb_activate.max_row)+1):
                if (n == wb_activate["A"+str(every_row)].value):
                    alreadyexist = True
                    break
            if alreadyexist == True:
                messagebox.showerror("DATA", "DATA ALREADY EXIST")
                root2.iconify()
            else:
                lastrow = str(wb_activate.max_row + 1)
                wb_activate['A'+ lastrow] = n
                wb_activate['B'+ lastrow] = a
                wb_activate['C'+ lastrow] = p
                wb_activate['D'+ lastrow] = l
                wb.save('library_data.xlsx')
                refresh_data(tv4)
                refresh_data(tv1)
        else:
            messagebox.showerror("DATA", "INCOMPLETE DATA")
            root2.iconify()
    def refresh_data(tree):
        tree.delete(*tree.get_children())
        data = update_data()
        for item in data:
            tree.insert('', 'end', values=item)
    def update_data():
        update_value = list()
        for each_cell in range(2, (wb_activate.max_row)+1):
            update_value.append([wb_activate['A'+str(each_cell)].value,wb_activate['B'+str(each_cell)].value,wb_activate['C'+str(each_cell)].value,wb_activate['D'+str(each_cell)].value])
        return update_value
    def remove():
        Found = False
        un = name_ent.get()
        for each_row in range(2, (wb_activate.max_row)+1):
            if (un == wb_activate["A"+str(each_row)].value):
                Found = True
                cell_address = each_row
                break
        if Found == True:
            wb_activate.delete_rows(cell_address)
            messagebox.showinfo("Deleted", "Data Has Been Deleted")
            root2.iconify()
            clear()
            refresh_data(tv4)
            wb.save('library_data.xlsx')
    def on_tree_select(event):
        selected_item = tv4.selection()
        if selected_item:
            book_data = tv4.item(selected_item)["values"]
            name_ent.delete(0, END)
            author_ent.delete(0, END)
            publishdate_ent.delete(0, END)
            location_ent.delete(0, END)
            code.delete(0, END)
            name_ent.insert(0, book_data[0])
            author_ent.insert(0, book_data[1])
            publishdate_ent.insert(0, book_data[2])
            location_ent.insert(0, book_data[3])
            for each_cell in range(2, (wb_activate.max_row)+1):
                if (name_ent.get() == wb_activate["A"+str(each_cell)].value):
                    cell_address = str(each_cell)
                    code.insert(0,wb_activate['F'+cell_address].value)
                    break
    def clear():
        name_ent.delete(0, END)
        author_ent.delete(0, END)
        publishdate_ent.delete(0, END)
        location_ent.delete(0, END)
        search_ent.delete(0, END)
        code.delete(0, END)
    def edit():
        n = name_ent.get()
        a = author_ent.get()
        pd = publishdate_ent.get()
        l = location_ent.get()
        c = code.get()
        if n and a and pd and l:
            for each_cell in range(2, (wb_activate.max_row)+1):
                if (c == wb_activate["F"+str(each_cell)].value):
                    cell_address = str(each_cell)
                    wb_activate['A'+ cell_address] = n
                    wb_activate['B'+ cell_address] = a
                    wb_activate['C'+ cell_address] = pd
                    wb_activate['D'+ cell_address] = l
                    wb.save('library_data.xlsx')
                    messagebox.showinfo("Edit", "Data Has Been Edited Successfully")
                    root2.iconify()
                    refresh_data(tv4)
                    refresh_data(tv1)
                    clear()
                    break
        else:
            messagebox.showerror("Edit Form", "No Book To Edit, \nEnter What Book You Like to Edit\nor Your Entry is Incomplete")
            root2.iconify()
    def lmview_data():
        global tv4
        tv4 = ttk.Treeview(lmSomeFrame, columns=("Name", "Author", "Publish Date","Location"), show="headings")
        tv4.heading("Name", text="Name", anchor=CENTER)
        tv4.heading("Author", text="Author", anchor=CENTER)
        tv4.heading("Publish Date", text="Publish Date", anchor=CENTER)
        tv4.heading("Location", text="Location", anchor=CENTER)
        for each_cell in range(2, (wb_activate.max_row)+1):
            tv4.insert(parent='', index="end", values=(wb_activate['A'+str(each_cell)].value, wb_activate['B'+str(each_cell)].value, wb_activate['C'+str(each_cell)].value, wb_activate['D'+str(each_cell)].value))
        tv4.column("Name", width=200)
        tv4.column("Author", width=150)
        tv4.column("Publish Date", width=100)
        tv4.column("Location", width=100)
        tv4.bind("<<TreeviewSelect>>", on_tree_select)
        tree_scroll = ttk.Scrollbar(lmSomeFrame, orient="vertical",command=tv4.yview)
        tv4.configure(yscrollcommand=tree_scroll.set)
        tv4.config(height=28)
        tv4.grid(row=0, column=0, sticky="nsew")
        tree_scroll.grid(row=0, column=1, sticky="ns")
    lmview_data()
    def lmsearch_library():
        query = search_ent.get().strip().lower()
        tv4.delete(*tv4.get_children())
        for row in wb_activate.iter_rows(min_row=2, values_only=True):
            if any(query in str(cell).lower() for cell in row):
                tv4.insert("", tk.END, values=row)
    name_lbl.grid(row=0, column=0, columnspan=2, sticky="ws", padx=10, pady=5)
    lmSomeFrame.grid(row=0, column=3, rowspan=20, padx=10, pady=10)
    name_ent.grid(row=1, column=0, columnspan=2, padx=10, ipady=3, sticky=N)
    author_lbl.grid(row=2, column=0, columnspan=2, sticky="ws", padx=10, pady=5)
    author_ent.grid(row=3, column=0, columnspan=2, padx=10, ipady=3, sticky=N)
    publishdate_lbl.grid(row=4, column=0, columnspan=2, sticky="ws", padx=10, pady=5)
    publishdate_ent.grid(row=5, column=0, columnspan=2, padx=10, ipady=3, sticky=N)
    location_lbl.grid(row=6, column=0, columnspan=2, sticky="ws", padx=10, pady=5)
    location_ent.grid(row=7, column=0, columnspan=2, padx=10, ipady=3, sticky=N)
    save_btn.grid(row=8, column=0, padx=5, pady=5)
    Remove_btn.grid(row=8, column=1, padx=5, pady=5)
    Edit_btn.grid(row=9, column=0, padx=5)
    clear_btn.grid(row=9, column=1, padx=5)
    search_ent.grid(row=10, column=0, columnspan=2, padx=10, ipady=3, sticky=S)
    search_btn.grid(row=11, column=0, padx=10, pady=5, sticky=N)
    back_btn.grid(row=11, column=1, padx=10, pady=5, sticky=N)
    #profile page
    pbackground = PhotoImage(file="profile bg.png")
    pbg = Label(profile_frame, image=pbackground, border=0)
    pframe = Frame(profile_frame, bg="#1A1C22")
    pframe2 = Frame(profile_frame, bg="#F0C39A")
    pframe3 = Frame(profile_frame, bg="#1A1C22")
    imagelist = ["logo.png"]
    randomimage = random.choice(imagelist)
    imagechoice = PhotoImage(file=randomimage)
    pprofile = Label(pframe2, image=imagechoice)
    def underline_text(widget):
        underline_font = font.Font(font="Garamond")
        underline_font.configure(underline=True)
        widget.configure(font=underline_font)
    changebtn1 = Button(pframe3, text="Change Profile Information", border=0, bg="#1A1C22", fg="#F0C39A", command=lambda:editprofile())
    changebtn2 = Button(pframe3, text="Change Password", border=0, bg="#1A1C22", fg="#F0C39A", command=lambda:editpassword())
    underline_text(changebtn1)
    underline_text(changebtn2)
    def profilerefresh():
        pname.grid_forget()
        page.grid_forget()
        psex.grid_forget()
        pphone.grid_forget()
        paddress.grid_forget()
        profileinfo()
    def profileinfo():
        global pname
        global page
        global psex
        global pphone
        global paddress
        nameval = Entry(profile_frame)
        ageval = Entry(profile_frame)
        phoneval = Entry(profile_frame)
        addressval = Entry(profile_frame)
        sexval = Entry(profile_frame)
        if mFound == True:
            nameval.insert(0,mm_activate['A'+cell_address].value)
            ageval.insert(0,mm_activate['B'+cell_address].value)
            phoneval.insert(0,mm_activate['C'+cell_address].value)
            addressval.insert(0,mm_activate['D'+cell_address].value)
            sexval.insert(0,mm_activate['E'+cell_address].value)
        if eFound == True:
            nameval.insert(0,em_activate['B'+cell_address2].value)
            ageval.insert(0,em_activate['C'+cell_address2].value)
            phoneval.insert(0,em_activate['D'+cell_address2].value)
            addressval.insert(0,em_activate['E'+cell_address2].value)
            sexval.insert(0,em_activate['F'+cell_address2].value)
        sname = nameval.get().strip()
        sage = ageval.get().strip() + " years old"
        sphone = phoneval.get().strip()
        saddress = addressval.get().strip()
        ssex = sexval.get().strip()
        pname = Label(pframe, text=sname, bg="#1A1C22", fg="white", font=("Garamond Bold", 35))
        page = Label(pframe, text=sage, bg="#1A1C22", fg="white", font=("Garamond", 25))
        pphone = Label(pframe, text=sphone, bg="#1A1C22", fg="white", font=("Garamond", 25))
        psex = Label(pframe, text=ssex, bg="#1A1C22", fg="white", font=("Garamond", 25))
        paddress = Label(pframe, text=saddress, bg="#1A1C22", fg="white", font=("Garamond", 20))
        pname.grid(row=0, column=0, sticky=W)
        page.grid(row=1, column=0, sticky=W)
        psex.grid(row=2, column=0, sticky=W)
        pphone.grid(row=3, column=0, sticky=W)
        paddress.grid(row=4, column=0, sticky=W)
    profileinfo()
    def editprofile():
        editwindow = Toplevel()
        editwindow.title("Edit Profile Info")
        editwindow.geometry("400x600")
        editwindow.resizable("False","False")
        bg1 = PhotoImage(file="bg.png")
        ebg = Label(editwindow, image=bg1)
        ebg.place(x=0, y=0, relwidth=1, relheight=1)
        emiddleframe = Frame(editwindow, bg="#1C2834")
        layout7 = Label(emiddleframe, bg="#1C2834")
        lblr_reg = Label(emiddleframe, text="Edit Profile", bg="#1C2834", fg="white", font=("High Tower Text", 30))
        lblr_name = Label(emiddleframe, text="Name:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
        ente_name = Entry(emiddleframe, width=30)
        lblr_age = Label(emiddleframe, text="Age:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
        ente_age = Entry(emiddleframe, width=30)
        lblr_gender = Label(emiddleframe,text="Gender:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
        esselect=StringVar()
        chk_male=Radiobutton(emiddleframe, text="Male",value="Male", bg="#1C2834", fg="#F0C39A", font=("High Tower Text", 12),variable=esselect)
        chk_female=Radiobutton(emiddleframe, text="Female",value="Female", bg="#1C2834", fg="#F0C39A", font=("High Tower Text", 12),variable=esselect)
        lblr_cont = Label(emiddleframe, text="Phone:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
        ente_cont = Entry(emiddleframe, width=30)
        lblr_add = Label(emiddleframe, text="Address:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
        ente_add = Text(emiddleframe, height= 5,width=23)
        update_profile = Button(emiddleframe, width=10, text="Save Change", font=("High Tower Text", 12), command=lambda:changeprofile())
        if mFound == True:
            ente_name.insert(0,mm_activate['A'+cell_address].value)
            ente_age.insert(0,mm_activate['B'+cell_address].value)
            ente_cont.insert(0,mm_activate['C'+cell_address].value)
            ente_add.insert(1.0,mm_activate['D'+cell_address].value)
        if eFound == True:
            ente_name.insert(0,em_activate['B'+cell_address2].value)
            ente_age.insert(0,em_activate['C'+cell_address2].value)
            ente_cont.insert(0,em_activate['D'+cell_address2].value)
            ente_add.insert(1.0,em_activate['E'+cell_address2].value)
        def changeprofile():
            if mFound == True:
                if ente_name.get() and ente_age.get() and ente_cont.get() and ente_add.get(1.0, END) and esselect.get():
                    if ente_age.get().isdigit():
                        if ente_cont.get().isdigit():
                            mm_activate['A'+str(each_cell)].value = ente_name.get()
                            mm_activate['B'+str(each_cell)].value = ente_age.get()
                            mm_activate['C'+str(each_cell)].value = ente_cont.get()
                            mm_activate['D'+str(each_cell)].value = ente_add.get(1.0, END)
                            mm_activate['E'+str(each_cell)].value = esselect.get()
                            mm.save('Member.xlsx')
                            messagebox.showinfo("Update","Data has been Updated")
                            root2.iconify()
                            profilerefresh()
                            editwindow.destroy()
                        else:
                            messagebox.showerror("Update","Invalid Phone Input")
                            root2.iconify()
                    else:
                        messagebox.showerror("Update","Age Input Incorrect")
                        root2.iconify()
                else:
                    messagebox.showerror("Update","Data input incomplete")
                    root2.iconify()
            if eFound == True:
                if ente_name.get() and ente_age.get() and ente_cont.get() and ente_add.get(1.0, END) and esselect.get():
                    if ente_age.get().isdigit():
                        if ente_cont.get().isdigit():
                            em_activate['B'+str(each_cell2)].value = ente_name.get()
                            em_activate['C'+str(each_cell2)].value = ente_age.get()
                            em_activate['D'+str(each_cell2)].value = ente_cont.get()
                            em_activate['E'+str(each_cell2)].value = ente_add.get(1.0, END)
                            em_activate['F'+str(each_cell2)].value = esselect.get()
                            em.save('Employee.xlsx')
                            messagebox.showinfo("Update","Data has been Updated")
                            root2.iconify()
                            profilerefresh()
                            editwindow.destroy()
                        else:
                            messagebox.showerror("Update","Invalid Phone Input")
                            root2.iconify()
                    else:
                        messagebox.showerror("Update","Age Input Incorrect")
                        root2.iconify()
                else:
                    messagebox.showerror("Update","Data input incomplete")
                    root2.iconify()
        emiddleframe.pack(expand=True)
        lblr_reg.grid(row=0, column=0, columnspan=3, pady=10)
        lblr_name.grid(row=2, column=0,padx=10, sticky=E)
        ente_name.grid(row=2, column=1,padx=5,columnspan=2, ipady=3, pady=2)
        lblr_age.grid(row=3, column=0,padx=10, sticky=E)
        ente_age.grid(row=3, column=1,padx=5,columnspan=2, ipady=3, pady=2)
        lblr_gender.grid(row=4, column=0,padx=10, sticky=E, pady=2)
        chk_male.grid(row=4, column=1,padx=5)
        chk_female.grid(row=4, column=2,padx=5)
        lblr_cont.grid(row=6, column=0,padx=10, sticky=E)
        ente_cont.grid(row=6, column=1,padx=5,columnspan=2, ipady=3, pady=2)
        lblr_add.grid(row=7, column=0,padx=10, sticky=E)
        ente_add.grid(row=7, column=1,padx=5,columnspan=2, ipady=3, pady=2)
        update_profile.grid(row=8, column=0,padx=5,columnspan=3, ipady=3, pady=2)
        layout7.grid(row=9, column=0, pady=80)
        editwindow.mainloop()
    def editpassword():
        editwindow2 = Toplevel()
        editwindow2.title("Edit User Password")
        editwindow2.resizable("False","False")
        editwindow2.configure(bg="#1C2834")
        oldpass = Label(editwindow2, text="Old Password:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
        oldpent = Entry(editwindow2, width=30)
        newpass = Label(editwindow2, text="New Password:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
        newpent = Entry(editwindow2, width=30)
        confpass = Label(editwindow2, text="Confirm Password:", bg="#1C2834", fg="white", font=("High Tower Text", 12))
        confpent = Entry(editwindow2, width=30)
        member_currentpass = Entry(editwindow2)
        employee_currentpass = Entry(editwindow2)
        if mFound == True:
            for each_cell in range(2, (mm_activate.max_row)+1):
                if (ent_user.get() == mm_activate['F'+str(each_cell)].value):
                    cell_address = str(each_cell)
                    member_currentpass.insert(0,mm_activate['G'+cell_address].value)
        if eFound == True:
            for each_cell2 in range(2, (em_activate.max_row)+1):
                if (ent_user.get() == em_activate['G'+str(each_cell2)].value):
                    cell_address2 = str(each_cell2)
                    employee_currentpass.insert(0,em_activate['H'+cell_address2].value)
                    print(employee_currentpass.get())
        savepass = Button(editwindow2, text="Change Password", command=lambda:savepassword())
        def savepassword():
            mc = member_currentpass.get()
            ec = employee_currentpass.get()
            o = oldpent.get()
            n = newpent.get()
            c = confpent.get()
            if mFound == True:
                if o and n and c:
                    if o == mc:
                        if mc != n:
                            if n == c:
                                mm_activate['G'+str(each_cell)].value = n
                                em.save('Employee.xlsx')
                                messagebox.showinfo("Update","New Password Has Been Saved")
                                root2.iconify()
                                editwindow2.destroy()
                            else:
                                messagebox.showerror("Update","CONFIRMATION FAILED")
                                root2.iconify()
                        else:
                            messagebox.showerror("Update","NEW PASSWORD ARE THE SAME AS THE OLD PASSWORD")
                            root2.iconify()
                    else:
                        messagebox.showerror("Update","OLD PASSWORD DOES NOT MATCH YOU CURRENT PASSWORD")
                        root2.iconify()
                else:
                    messagebox.showerror("Update","DATA INPUT INCOMPLETE")
                    root2.iconify()
            if eFound == True:
                if o and n and c:
                    if o == ec:
                        if n == c:
                            em_activate['H'+str(each_cell2)].value = n
                            em.save('Employee.xlsx')
                            messagebox.showinfo("Update","New Password Has Been Saved")
                            root2.iconify()
                            editwindow2.destroy()
                        else:
                            messagebox.showerror("Update","CONFIRMATION FAILED")
                            root2.iconify()
                    else:
                        messagebox.showerror("Update","OLD PASSWORD DOES NOT MATCH YOU ORIGINAL PASSWORD")
                        root2.iconify()
                else:
                    messagebox.showerror("Update","DATA INPUT INCOMPLETE")
                    root2.iconify()
        oldpass.grid(row=0, column=0, sticky=E, padx=3, pady=3)
        oldpent.grid(row=0, column=1, padx=3, pady=3)
        newpass.grid(row=1, column=0, sticky=E, padx=3, pady=3)
        newpent.grid(row=1, column=1, padx=3, pady=3)
        confpass.grid(row=2, column=0, sticky=E, padx=3, pady=3)
        confpent.grid(row=2, column=1, padx=3, pady=3)
        savepass.grid(row=3, column=0, columnspan=2, padx=3, pady=3)
        editwindow2.mainloop()
    #profile layout
    pbg.pack()
    pframe.place(x=200, y=150)
    pframe2.place(x=20, y=80)
    pframe3.place(x=690, y=450)
    pprofile.pack(padx=5, pady=5)
    changebtn1.grid(row=5, column=0, sticky=E)
    changebtn2.grid(row=6, column=0, sticky=E)
    #help_frame
    #page1
    frame1 = Frame(help_frame, bg="#1A1C22")
    frame1.pack()
    frame12 = Frame(frame1, bg="#1A1C22")
    ph1 = PhotoImage(file="helppage1.png")
    page1 = Label(frame12, image=ph1, border=0)
    frame12.grid(row=1, column=0)
    page1.pack(padx=100)
    frame11 = Frame(frame1, bg="#1A1C22")
    nextbtn1 = Button(frame11, text="Next", width=8 , bg="#F0C39A", fg="#1A1C22", font=("Garamond", 12), command=lambda:next1())
    page1layout = Label(frame11, width=100, bg="#1A1C22")
    frame11.grid(row=2, column=0)
    previousbtn1 = Button(frame11, text="Previous", width=8, bg="#F0C39A", fg="#1A1C22", font=("Garamond", 12))
    page1layout.grid(row=0, column=1)
    previousbtn1.grid(row=0, column=0, sticky=E, pady=10)
    nextbtn1.grid(row=0, column=2, sticky=W, pady=10)
    frame13 = Frame(frame1, bg="#1A1C22")
    helplbl = Label(frame13, text="Find Books", bg="#1A1C22", fg="#F0C39A", font=("Garamond", 50))
    frame13.grid(row=0, column=0)
    helplbl.pack()
    cat1 = Label(frame12, bg="White", text="First nya, Let's go to the library \nso nya can checks on the books, Nya!", font=("Garamond", 12))
    cat2 = Label(frame12, bg="White", text="Then type in the name of the \nbook you seek Nya!", font=("Garamond", 12))
    cat1.place(x=330, y=20)
    cat2.place(x=235, y=420)
    #page2
    frame2 = Frame(help_frame, bg="#1A1C22")
    frame23 = Frame(frame2, bg="#1A1C22")
    helplbl = Label(frame23, text="Find Books", bg="#1A1C22", fg="#F0C39A", font=("Garamond", 50))
    frame23.grid(row=0, column=0)
    helplbl.pack()
    frame21 = Frame(frame2, bg="#1A1C22")
    ph2 = PhotoImage(file="helppage2.png")
    page2 = Label(frame21, image=ph2, border=0)
    frame21.grid(row=1, column=0)
    page2.pack(padx=98)
    frame22 = Frame(frame2, bg="#1A1C22")
    nextbtn2 = Button(frame22, text="Next", width=8 , bg="#F0C39A", fg="#1A1C22", font=("Garamond", 12))
    previousbtn2 = Button(frame22, text="Previous", width=8, bg="#F0C39A", fg="#1A1C22", font=("Garamond", 12), command=lambda:previous2())
    page2layout = Label(frame22, width=100, bg="#1A1C22")
    frame22.grid(row=2, column=0)
    previousbtn2.grid(row=0, column=0, sticky=E, pady=10)
    page2layout.grid(row=0, column=1)
    nextbtn2.grid(row=0, column=2, sticky=W)
    cat3 = Label(frame21, bg="White", text="Nyext! let's press search and it \nshould show or narrow down the \nselections of the books!", font=("Garamond", 11))
    cat4 = Label(frame21, bg="White", text="And then check the location \nof nya book in the location \ncolumn nya!!", font=("Garamond", 12))
    cat3.place(x=350, y=150)
    cat4.place(x=180, y=385)
    def next1():
        frame1.pack_forget()
        frame2.pack()
    def previous2():
        frame2.pack_forget()
        frame1.pack()
    Tl.mainloop()
root2.mainloop()